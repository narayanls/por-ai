"""
Geração de planilhas formatadas (.ods e .xlsx) a partir de uma spec JSON.

O formato padrão é ODS (OpenDocument) — formato aberto, coerente com um
app livre e legível por LibreOffice, Excel e Google Sheets. XLSX continua
disponível para quem pedir explicitamente.

Os dois backends partem da MESMA especificação e da MESMA paleta, então
uma planilha gerada em ods e outra em xlsx saem visualmente idênticas.
Todas as chaves de "style" são opcionais e caem no tema padrão; cores
inválidas são ignoradas em vez de quebrar a geração, porque o valor vem
de um LLM e não dá para confiar.

Cada backend é opcional: se openpyxl faltar, ods continua funcionando, e
vice-versa. O chamador consulta ``available_formats()`` antes de decidir.
"""

from __future__ import annotations

import re
from typing import Any, Dict, List, Sequence, Tuple

# ── Backends opcionais ───────────────────────────────────────────────────────

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    XLSX_BACKEND = True
except ImportError:
    XLSX_BACKEND = False

try:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.style import (
        ParagraphProperties,
        Style,
        TableCellProperties,
        TableColumnProperties,
        TextProperties,
    )
    from odf.table import Table, TableCell, TableColumn, TableRow
    from odf.text import P
    ODS_BACKEND = True
except ImportError:
    ODS_BACKEND = False


# ── Estilo de fallback ───────────────────────────────────────────────────────
#
# NÃO é um "tema do app": é o que sobra quando o modelo omite a chave style
# ou manda hex inválido. Por isso é deliberadamente neutro — quem escolhe a
# paleta é o usuário, via pedido ("gruvbox", "cyberpunk", "everforest"), e o
# modelo traduz isso para o style. Um tema aqui viraria política implícita.
#
# Fundo claro de propósito: planilha costuma acabar impressa ou aberta em
# app de terceiro com fundo branco, onde um fallback escuro fica ilegível.

_DEFAULT_STYLE = {
    "header_bg": "3A4750",
    "header_fg": "FFFFFF",
    "row_bg": "FFFFFF",
    "alt_row_bg": "F2F4F5",
    "row_fg": "1F2933",
    "border": "D5DADE",
}

_RE_HEX = re.compile(r"[0-9A-F]{6}")
_RE_BAD_SHEET_NAME = re.compile(r"[\[\]:*?/\\]")

_MAX_WIDTH_CHARS = 50
_WIDTH_PADDING = 6


class SheetSpecError(ValueError):
    """Especificação inválida (JSON fora do formato esperado)."""


# Alias mantido para não quebrar imports antigos de core.xlsx_gen.
XlsxSpecError = SheetSpecError


def available_formats() -> Tuple[str, ...]:
    """Formatos que podem ser gerados nesta instalação, em ordem de preferência."""
    formats = []
    if ODS_BACKEND:
        formats.append("ods")
    if XLSX_BACKEND:
        formats.append("xlsx")
    return tuple(formats)


def _hex_color(value: Any, fallback: str) -> str:
    """Normaliza '#RRGGBB' → 'RRGGBB'; devolve o padrão se não for válido."""
    if not isinstance(value, str):
        return fallback
    cleaned = value.lstrip("#").upper()
    return cleaned if _RE_HEX.fullmatch(cleaned) else fallback


def _parse_spec(spec: Dict[str, Any]) -> Tuple[List[str], List[list], Dict[str, str], str]:
    """Valida e normaliza a spec, compartilhado pelos dois backends."""
    columns = spec.get("columns")
    rows = spec.get("rows")
    if not isinstance(columns, list) or not columns:
        raise SheetSpecError("chave 'columns' ausente ou vazia")
    if not isinstance(rows, list):
        raise SheetSpecError("chave 'rows' ausente ou não é lista")

    columns = [str(c) for c in columns]
    rows = [list(r) for r in rows if isinstance(r, (list, tuple))]

    style = dict(_DEFAULT_STYLE)
    if isinstance(spec.get("style"), dict):
        style.update(spec["style"])
    style = {key: _hex_color(style.get(key), default)
             for key, default in _DEFAULT_STYLE.items()}

    sheet_name = spec.get("sheet_name")
    sheet_name = str(sheet_name)[:31] if sheet_name else "Planilha"
    sheet_name = _RE_BAD_SHEET_NAME.sub("-", sheet_name) or "Planilha"

    return columns, rows, style, sheet_name


def _column_widths(columns: Sequence[str], rows: Sequence[list]) -> List[int]:
    """Largura em caracteres por coluna, com teto pra não estourar a janela."""
    widths = []
    for index, title in enumerate(columns):
        widest = len(title)
        for row in rows:
            if len(row) > index:
                widest = max(widest, len(str(row[index])))
        widths.append(min(widest + _WIDTH_PADDING, _MAX_WIDTH_CHARS))
    return widths


def build_sheet(spec: Dict[str, Any], path: str, fmt: str = "") -> str:
    """Escreve a planilha descrita por ``spec`` em ``path`` e devolve o caminho.

    ``fmt`` é 'ods' ou 'xlsx'; se vazio, é deduzido da extensão de ``path``.
    Levanta ``SheetSpecError`` se a spec não tiver colunas/linhas utilizáveis
    ou se o backend do formato pedido não estiver instalado — nesse caso o
    chamador deve preservar o bloco original na resposta em vez de exibir um
    link para uma planilha vazia.
    """
    fmt = (fmt or path.rsplit(".", 1)[-1]).lower()
    if fmt == "ods":
        if not ODS_BACKEND:
            raise SheetSpecError("odfpy não instalado")
        return _build_ods(spec, path)
    if fmt in ("xlsx", "xls"):
        if not XLSX_BACKEND:
            raise SheetSpecError("openpyxl não instalado")
        return _build_xlsx(spec, path)
    raise SheetSpecError(f"formato não suportado: {fmt}")


# Nome antigo, mantido para compatibilidade.
def build_xlsx(spec: Dict[str, Any], path: str) -> str:
    return build_sheet(spec, path, "xlsx")


# ── Backend XLSX (openpyxl) ──────────────────────────────────────────────────

def _build_xlsx(spec: Dict[str, Any], path: str) -> str:
    columns, rows, style, sheet_name = _parse_spec(spec)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    edge = Side(style="thin", color=style["border"])
    border = Border(left=edge, right=edge, top=edge, bottom=edge)

    header_font = Font(bold=True, size=12, color=style["header_fg"])
    header_fill = PatternFill("solid", fgColor=style["header_bg"])

    ws.append(columns)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    row_font = Font(color=style["row_fg"])
    fills = (
        PatternFill("solid", fgColor=style["row_bg"]),
        PatternFill("solid", fgColor=style["alt_row_bg"]),
    )

    for index, row in enumerate(rows):
        ws.append(row)
        fill = fills[index % 2]
        for cell in ws[ws.max_row]:
            cell.font = row_font
            cell.fill = fill
            cell.border = border

    for index, width in enumerate(_column_widths(columns, rows), start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
    return path


# ── Backend ODS (odfpy) ──────────────────────────────────────────────────────
#
# O ODF não tem "largura de coluna em caracteres" como o xlsx: a medida é
# física. 0.22cm por caractere é a aproximação que mais se aproxima do
# resultado do openpyxl na fonte padrão do LibreOffice.

_CM_PER_CHAR = 0.22


def _ods_cell(value: Any, style) -> "TableCell":
    """Célula tipada: números viram float de verdade, não texto.

    Sem isso a planilha abre com tudo alinhado à esquerda e o usuário não
    consegue somar uma coluna. bool vem antes de int de propósito — em
    Python ``bool`` é subclasse de ``int``.
    """
    if isinstance(value, bool):
        cell = TableCell(valuetype="boolean", booleanvalue="true" if value else "false")
        text = "VERDADEIRO" if value else "FALSO"
    elif isinstance(value, (int, float)):
        cell = TableCell(valuetype="float", value=str(value))
        text = str(value)
    elif value is None:
        cell = TableCell(valuetype="string")
        text = ""
    else:
        cell = TableCell(valuetype="string")
        text = str(value)
    cell.setAttribute("stylename", style)
    cell.addElement(P(text=text))
    return cell


def _build_ods(spec: Dict[str, Any], path: str) -> str:
    columns, rows, style, sheet_name = _parse_spec(spec)

    doc = OpenDocumentSpreadsheet()

    def _cell_style(name: str, bg: str, fg: str, bold: bool, center: bool):
        cell_style = Style(name=name, family="table-cell")
        cell_style.addElement(TableCellProperties(
            backgroundcolor=f"#{bg}",
            border=f"0.5pt solid #{style['border']}",
        ))
        cell_style.addElement(TextProperties(
            color=f"#{fg}",
            fontweight="bold" if bold else "normal",
            fontsize="12pt" if bold else "10pt",
        ))
        if center:
            cell_style.addElement(ParagraphProperties(textalign="center"))
        doc.automaticstyles.addElement(cell_style)
        return cell_style

    header_style = _cell_style("hdr", style["header_bg"], style["header_fg"], True, True)
    body_styles = (
        _cell_style("bd0", style["row_bg"], style["row_fg"], False, False),
        _cell_style("bd1", style["alt_row_bg"], style["row_fg"], False, False),
    )

    table = Table(name=sheet_name)

    for width in _column_widths(columns, rows):
        col_style = Style(name=f"co{width}", family="table-column")
        col_style.addElement(TableColumnProperties(
            columnwidth=f"{width * _CM_PER_CHAR:.2f}cm"
        ))
        doc.automaticstyles.addElement(col_style)
        table.addElement(TableColumn(stylename=col_style))

    header_row = TableRow()
    for title in columns:
        header_row.addElement(_ods_cell(title, header_style))
    table.addElement(header_row)

    for index, row in enumerate(rows):
        table_row = TableRow()
        cell_style = body_styles[index % 2]
        for value in row:
            table_row.addElement(_ods_cell(value, cell_style))
        # Preenche o rabo curto pra zebra não ficar com buraco branco.
        for _ in range(len(columns) - len(row)):
            table_row.addElement(_ods_cell("", cell_style))
        table.addElement(table_row)

    doc.spreadsheet.addElement(table)
    doc.save(path)
    return path