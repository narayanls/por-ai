"""
Geração de planilhas formatadas (.ods e .xlsx) a partir de uma spec JSON.

O formato padrão é ODS (OpenDocument) — formato aberto, coerente com um
app livre e legível por LibreOffice, Excel e Google Sheets. XLSX continua
disponível para quem pedir explicitamente.

Os dois backends partem da MESMA especificação e da MESMA paleta, então
uma planilha gerada em ods e outra em xlsx saem visualmente idênticas.
Todas as chaves de "style" são opcionais e caem no tema padrão; cores
inválidas são ignoradas em vez de quebrar a geração, porque o valor vem
de um LLM e não dá para confiar.

Cada backend é opcional: se openpyxl faltar, ods continua funcionando, e
vice-versa. O chamador consulta ``available_formats()`` antes de decidir.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any, Dict, List, Optional, Sequence, Tuple

logger = logging.getLogger(__name__)

# ── Backends opcionais ───────────────────────────────────────────────────────

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    XLSX_BACKEND = True
except ImportError:
    XLSX_BACKEND = False

try:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.style import (
        ParagraphProperties,
        Style,
        TableCellProperties,
        TableColumnProperties,
        TextProperties,
    )
    from odf.table import Table, TableCell, TableColumn, TableRow
    from odf.text import P
    from odf.draw import Frame as DrawFrame, Image as DrawImage
    ODS_BACKEND = True
except ImportError:
    ODS_BACKEND = False

# Gráfico é opcional e independente dos backends de planilha: sem
# matplotlib, os dois formatos continuam gerando normalmente, só sem a
# imagem do gráfico embutida.
from core import chart_gen


# ── Estilo de fallback ───────────────────────────────────────────────────────
#
# NÃO é um "tema do app": é o que sobra quando o modelo omite a chave style
# ou manda hex inválido. Por isso é deliberadamente neutro — quem escolhe a
# paleta é o usuário, via pedido ("gruvbox", "cyberpunk", "everforest"), e o
# modelo traduz isso para o style. Um tema aqui viraria política implícita.
#
# Fundo claro de propósito: planilha costuma acabar impressa ou aberta em
# app de terceiro com fundo branco, onde um fallback escuro fica ilegível.

_DEFAULT_STYLE = {
    "header_bg": "3A4750",
    "header_fg": "FFFFFF",
    "row_bg": "FFFFFF",
    "alt_row_bg": "F2F4F5",
    "row_fg": "1F2933",
    "border": "D5DADE",
}

_RE_HEX = re.compile(r"[0-9A-F]{6}")
_RE_BAD_SHEET_NAME = re.compile(r"[\[\]:*?/\\]")

_MAX_WIDTH_CHARS = 50
_WIDTH_PADDING = 6


class SheetSpecError(ValueError):
    """Especificação inválida (JSON fora do formato esperado)."""


# Alias mantido para não quebrar imports antigos de core.xlsx_gen.
XlsxSpecError = SheetSpecError


def available_formats() -> Tuple[str, ...]:
    """Formatos que podem ser gerados nesta instalação, em ordem de preferência."""
    formats = []
    if ODS_BACKEND:
        formats.append("ods")
    if XLSX_BACKEND:
        formats.append("xlsx")
    return tuple(formats)


def chart_available() -> bool:
    """True se matplotlib está instalado e gráficos podem ser embutidos."""
    return chart_gen.CHART_BACKEND


def capabilities_prompt() -> str:
    """Texto a acrescentar ao system prompt, descrevendo o que este
    ambiente sabe gerar agora mesmo.

    É montado em runtime (não é texto fixo em config.py) de propósito:
    reflete exatamente os backends instalados nesta máquina, e o usuário
    pode editar o system_prompt em Preferências sem correr o risco de
    apagar ou desalinhar as instruções de formato. Se nenhum backend de
    planilha estiver disponível, devolve string vazia — nesse caso o
    modelo nunca é instruído a produzir blocos que ninguém vai converter.
    """
    formats = available_formats()
    if not formats:
        return ""

    preferred = formats[0]  # ods, se disponível — ver comentário no topo do arquivo
    lines = [
        "Ao gerar uma planilha para o usuário, produza um bloco de código "
        f"cercado com ```{preferred} (ou ```xlsx apenas se o usuário pedir "
        "Excel explicitamente) contendo APENAS um JSON válido, sem texto "
        "fora da cerca, no formato:",
        '{"sheet_name": "...", "columns": ["Col1", "Col2", ...], '
        '"rows": [[v1, v2, ...], ...], "style": {"header_bg": "RRGGBB", ...}}',
        "'style' é opcional (cores em hex sem #: header_bg, header_fg, "
        "row_bg, alt_row_bg, row_fg, border). Números devem ir como number "
        "JSON, não como string, para a planilha reconhecer a coluna como "
        "numérica.",
    ]
    if "xlsx" not in formats:
        lines.append("Não gere blocos ```xlsx nesta instalação: apenas ```ods está disponível.")
    elif "ods" not in formats:
        lines.append("Não gere blocos ```ods nesta instalação: apenas ```xlsx está disponível.")

    if chart_available():
        lines.append(
            "Se o usuário pedir um gráfico junto com a planilha (ou dados "
            "que claramente pedem visualização), acrescente uma chave "
            '"chart" no MESMO JSON — não crie um bloco separado nem repita '
            "os dados. Formato:"
        )
        lines.append(
            '"chart": {"type": "bar" | "line" | "pie" | "scatter", '
            '"title": "...", "category_column": "nome exato de uma coluna '
            'em columns", "value_columns": ["nome de coluna", ...], '
            '"x_label": "...", "y_label": "..."}'
        )
        lines.append(
            "category_column e value_columns DEVEM ser nomes que já "
            "existem em 'columns' — nunca invente uma coluna nova só para "
            "o gráfico. Para 'scatter', use 'x_column' no lugar de "
            "'category_column'. Para 'pie', informe só uma coluna em "
            "value_columns (as demais são ignoradas)."
        )
    else:
        lines.append(
            "Gráficos NÃO estão disponíveis nesta instalação (dependência "
            "ausente): não inclua a chave 'chart' na planilha, e avise o "
            "usuário que a instalação atual não gera gráficos se ele pedir um."
        )

    return "\n".join(lines)


def _hex_color(value: Any, fallback: str) -> str:
    """Normaliza '#RRGGBB' → 'RRGGBB'; devolve o padrão se não for válido."""
    if not isinstance(value, str):
        return fallback
    cleaned = value.lstrip("#").upper()
    return cleaned if _RE_HEX.fullmatch(cleaned) else fallback


def _parse_spec(
    spec: Dict[str, Any],
) -> Tuple[List[str], List[list], Dict[str, str], str, Optional[Dict[str, Any]]]:
    """Valida e normaliza a spec, compartilhado pelos dois backends."""
    columns = spec.get("columns")
    rows = spec.get("rows")
    if not isinstance(columns, list) or not columns:
        raise SheetSpecError("chave 'columns' ausente ou vazia")
    if not isinstance(rows, list):
        raise SheetSpecError("chave 'rows' ausente ou não é lista")

    columns = [str(c) for c in columns]
    rows = [list(r) for r in rows if isinstance(r, (list, tuple))]

    style = dict(_DEFAULT_STYLE)
    if isinstance(spec.get("style"), dict):
        style.update(spec["style"])
    style = {key: _hex_color(style.get(key), default)
             for key, default in _DEFAULT_STYLE.items()}

    sheet_name = spec.get("sheet_name")
    sheet_name = str(sheet_name)[:31] if sheet_name else "Planilha"
    sheet_name = _RE_BAD_SHEET_NAME.sub("-", sheet_name) or "Planilha"

    # Chave opcional: só é usada se vier um dict válido. Qualquer outra
    # coisa (string, lista, ausente) é tratada como "sem gráfico" em vez
    # de erro — a planilha não pode falhar por causa de um extra opcional.
    chart_spec = spec.get("chart")
    chart_spec = chart_spec if isinstance(chart_spec, dict) else None

    return columns, rows, style, sheet_name, chart_spec


def _column_widths(columns: Sequence[str], rows: Sequence[list]) -> List[int]:
    """Largura em caracteres por coluna, com teto pra não estourar a janela."""
    widths = []
    for index, title in enumerate(columns):
        widest = len(title)
        for row in rows:
            if len(row) > index:
                widest = max(widest, len(str(row[index])))
        widths.append(min(widest + _WIDTH_PADDING, _MAX_WIDTH_CHARS))
    return widths


def build_sheet(spec: Dict[str, Any], path: str, fmt: str = "") -> str:
    """Escreve a planilha descrita por ``spec`` em ``path`` e devolve o caminho.

    ``fmt`` é 'ods' ou 'xlsx'; se vazio, é deduzido da extensão de ``path``.
    Levanta ``SheetSpecError`` se a spec não tiver colunas/linhas utilizáveis
    ou se o backend do formato pedido não estiver instalado — nesse caso o
    chamador deve preservar o bloco original na resposta em vez de exibir um
    link para uma planilha vazia.
    """
    fmt = (fmt or path.rsplit(".", 1)[-1]).lower()
    if fmt == "ods":
        if not ODS_BACKEND:
            raise SheetSpecError("odfpy não instalado")
        return _build_ods(spec, path)
    if fmt in ("xlsx", "xls"):
        if not XLSX_BACKEND:
            raise SheetSpecError("openpyxl não instalado")
        return _build_xlsx(spec, path)
    raise SheetSpecError(f"formato não suportado: {fmt}")


# Nome antigo, mantido para compatibilidade.
def build_xlsx(spec: Dict[str, Any], path: str) -> str:
    return build_sheet(spec, path, "xlsx")


# ── Backend XLSX (openpyxl) ──────────────────────────────────────────────────

def _build_xlsx(spec: Dict[str, Any], path: str) -> str:
    columns, rows, style, sheet_name, chart_spec = _parse_spec(spec)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    edge = Side(style="thin", color=style["border"])
    border = Border(left=edge, right=edge, top=edge, bottom=edge)

    header_font = Font(bold=True, size=12, color=style["header_fg"])
    header_fill = PatternFill("solid", fgColor=style["header_bg"])

    ws.append(columns)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    row_font = Font(color=style["row_fg"])
    fills = (
        PatternFill("solid", fgColor=style["row_bg"]),
        PatternFill("solid", fgColor=style["alt_row_bg"]),
    )

    for index, row in enumerate(rows):
        ws.append(row)
        fill = fills[index % 2]
        for cell in ws[ws.max_row]:
            cell.font = row_font
            cell.fill = fill
            cell.border = border

    for index, width in enumerate(_column_widths(columns, rows), start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"

    if chart_spec is not None:
        _embed_chart_xlsx(ws, chart_spec, columns, rows, style, anchor_row=len(rows) + 3)

    wb.save(path)
    return path


def _embed_chart_xlsx(
    ws, chart_spec: Dict[str, Any], columns: List[str], rows: List[list],
    style: Dict[str, str], anchor_row: int,
) -> None:
    """Gera o PNG do gráfico e o embute abaixo da tabela. Falha em silêncio
    (log apenas) — um gráfico ruim não deve impedir a entrega da tabela."""
    png_bytes = chart_gen.render_chart(chart_spec, columns, rows, style)
    if not png_bytes:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
        image = XLImage(io.BytesIO(png_bytes))
        # Dimensões explícitas: openpyxl só precisaria do Pillow pra
        # descobrir isso sozinho a partir do PNG, e Pillow não é uma
        # dependência que queremos exigir só por causa do gráfico.
        image.width, image.height = 640, 384
        ws.add_image(image, f"A{anchor_row}")
    except Exception:  # pylint: disable=broad-except
        logger.exception("Falha ao embutir gráfico no xlsx")


# ── Backend ODS (odfpy) ──────────────────────────────────────────────────────
#
# O ODF não tem "largura de coluna em caracteres" como o xlsx: a medida é
# física. 0.22cm por caractere é a aproximação que mais se aproxima do
# resultado do openpyxl na fonte padrão do LibreOffice.

_CM_PER_CHAR = 0.22


def _ods_cell(value: Any, style) -> "TableCell":
    """Célula tipada: números viram float de verdade, não texto.

    Sem isso a planilha abre com tudo alinhado à esquerda e o usuário não
    consegue somar uma coluna. bool vem antes de int de propósito — em
    Python ``bool`` é subclasse de ``int``.
    """
    if isinstance(value, bool):
        cell = TableCell(valuetype="boolean", booleanvalue="true" if value else "false")
        text = "VERDADEIRO" if value else "FALSO"
    elif isinstance(value, (int, float)):
        cell = TableCell(valuetype="float", value=str(value))
        text = str(value)
    elif value is None:
        cell = TableCell(valuetype="string")
        text = ""
    else:
        cell = TableCell(valuetype="string")
        text = str(value)
    cell.setAttribute("stylename", style)
    cell.addElement(P(text=text))
    return cell


def _build_ods(spec: Dict[str, Any], path: str) -> str:
    columns, rows, style, sheet_name, chart_spec = _parse_spec(spec)

    doc = OpenDocumentSpreadsheet()

    def _cell_style(name: str, bg: str, fg: str, bold: bool, center: bool):
        cell_style = Style(name=name, family="table-cell")
        cell_style.addElement(TableCellProperties(
            backgroundcolor=f"#{bg}",
            border=f"0.5pt solid #{style['border']}",
        ))
        cell_style.addElement(TextProperties(
            color=f"#{fg}",
            fontweight="bold" if bold else "normal",
            fontsize="12pt" if bold else "10pt",
        ))
        if center:
            cell_style.addElement(ParagraphProperties(textalign="center"))
        doc.automaticstyles.addElement(cell_style)
        return cell_style

    header_style = _cell_style("hdr", style["header_bg"], style["header_fg"], True, True)
    body_styles = (
        _cell_style("bd0", style["row_bg"], style["row_fg"], False, False),
        _cell_style("bd1", style["alt_row_bg"], style["row_fg"], False, False),
    )

    table = Table(name=sheet_name)

    for width in _column_widths(columns, rows):
        col_style = Style(name=f"co{width}", family="table-column")
        col_style.addElement(TableColumnProperties(
            columnwidth=f"{width * _CM_PER_CHAR:.2f}cm"
        ))
        doc.automaticstyles.addElement(col_style)
        table.addElement(TableColumn(stylename=col_style))

    header_row = TableRow()
    for title in columns:
        header_row.addElement(_ods_cell(title, header_style))
    table.addElement(header_row)

    for index, row in enumerate(rows):
        table_row = TableRow()
        cell_style = body_styles[index % 2]
        for value in row:
            table_row.addElement(_ods_cell(value, cell_style))
        # Preenche o rabo curto pra zebra não ficar com buraco branco.
        for _ in range(len(columns) - len(row)):
            table_row.addElement(_ods_cell("", cell_style))
        table.addElement(table_row)

    if chart_spec is not None:
        _embed_chart_ods(doc, table, chart_spec, columns, rows, style)

    doc.spreadsheet.addElement(table)
    doc.save(path)
    return path


def _embed_chart_ods(
    doc: "OpenDocumentSpreadsheet", table: "Table", chart_spec: Dict[str, Any],
    columns: List[str], rows: List[list], style: Dict[str, str],
) -> None:
    """Gera o PNG do gráfico e o embute como imagem numa linha extra abaixo
    da tabela. Sem chart nativo do ODF (ver nota no topo do arquivo) — é
    uma imagem, igual ao xlsx, pra manter os dois backends consistentes."""
    png_bytes = chart_gen.render_chart(chart_spec, columns, rows, style)
    if not png_bytes:
        return
    try:
        href = doc.addPictureFromString(png_bytes, "image/png")
        frame = DrawFrame(width="16.9cm", height="10.1cm", anchortype="paragraph")
        frame.addElement(DrawImage(href=href))

        spacer_row = TableRow()
        for _ in columns:
            spacer_row.addElement(TableCell())
        table.addElement(spacer_row)

        chart_row = TableRow()
        chart_cell = TableCell()
        chart_cell.addElement(frame)
        chart_row.addElement(chart_cell)
        for _ in range(len(columns) - 1):
            chart_row.addElement(TableCell())
        table.addElement(chart_row)
    except Exception:  # pylint: disable=broad-except
        logger.exception("Falha ao embutir gráfico no ods")
