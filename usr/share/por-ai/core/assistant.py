"""
Coordenador de chat do POR.ai.

Formatos de texto suportados (enviados como bloco na mensagem):
  txt, md, rst, org, tex, csv, log, pdf, odt, ods, xlsx

Formatos de imagem suportados (enviados como base64 multimodal):
  jpg, jpeg, png, webp

Planilhas geradas pelo modelo: blocos ```ods e ```xlsx (JSON com dados e
estilo) viram planilhas formatadas; blocos ```csv viram .csv sem
formatação. Em todos os casos o bloco é trocado por um link file://
clicável. ODS é o formato pedido no system prompt; xlsx só quando o
usuário pede Excel explicitamente.
"""

from __future__ import annotations

import base64
import json
import logging
import mimetypes
import os
import re
import threading
import time
import uuid
from typing import Any, Callable, Dict, List, Optional, Tuple

from gi.repository import GLib

from core.config import Config
from core.openrouter import OpenRouterClient, OpenRouterError

# ── Dependências opcionais ────────────────────────────────────────────────────

try:
    from pypdf import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from odf.opendocument import load as odf_load
    from odf.text import P
    ODT_AVAILABLE = True
except ImportError:
    ODT_AVAILABLE = False

# openpyxl é tratado como opcional pelo mesmo motivo dos anteriores: se o
# pacote faltar na máquina do usuário, o app continua funcionando em vez de
# quebrar no import. Aqui a flag governa apenas a LEITURA de anexos .xlsx —
# a geração é decidida por `available_formats()`, que consulta os dois
# backends (odfpy e openpyxl) de forma independente.
try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# sheet_gen não tem dependência dura: cada backend se autodetecta lá dentro,
# então o import nunca falha por falta de odfpy ou openpyxl.
from core.sheet_gen import (
    SheetSpecError,
    available_formats,
    build_sheet,
    capabilities_prompt,
)


# ── Extensões suportadas ──────────────────────────────────────────────────────

_TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".rst", ".org",
    ".tex", ".csv", ".log",
}
_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}

logger = logging.getLogger(__name__)


class ChatAssistant:
    """Envia conversas ao OpenRouter sem travar a interface."""

    # Tempo de vida do cache do catálogo de modelos (evita bater na API do
    # OpenRouter a cada mensagem só pra saber a janela de contexto).
    _MODEL_CACHE_TTL = 3600.0

    def __init__(self, config: Config) -> None:
        self.config = config
        self._lock = threading.RLock()
        self._inflight = False
        self._cancel = threading.Event()
        self._model_cache: Dict[str, Dict[str, Any]] = {}
        self._model_cache_time: float = 0.0
        self._model_cache_lock = threading.Lock()

    # ------------------------------------------------------------------ #
    # Estado                                                               #
    # ------------------------------------------------------------------ #

    def is_busy(self) -> bool:
        with self._lock:
            return self._inflight

    def cancel(self) -> None:
        self._cancel.set()

    @staticmethod
    def _inject_capabilities(messages: List[Any]) -> List[Any]:
        """Acrescenta as instruções de formato de planilha/gráfico (ver
        sheet_gen.capabilities_prompt()) à mensagem de sistema, sem tocar
        no ``system_prompt`` que o usuário edita em Preferências.

        Gerado a cada envio (não uma vez só) porque reflete os backends
        realmente instalados agora — o texto muda se o usuário instalar ou
        remover odfpy/openpyxl/matplotlib sem reiniciar o app. Devolve uma
        lista NOVA: a lista original pode ser reusada pelo chamador (ex.:
        reenviar o histórico após editar uma mensagem) e não deve carregar
        esse texto interno.
        """
        addendum = capabilities_prompt()
        if not addendum:
            return messages

        result = list(messages)
        for index, message in enumerate(result):
            if isinstance(message, dict) and message.get("role") == "system":
                content = message.get("content", "")
                if isinstance(content, str):
                    merged = dict(message)
                    merged["content"] = f"{content}\n\n{addendum}" if content else addendum
                    result[index] = merged
                    return result
                # Mensagem de sistema em formato multimodal (lista de
                # blocos): não deveria acontecer na prática, mas se
                # acontecer é mais seguro deixar intacta a acrescentar uma
                # nova mensagem de sistema do que arriscar corromper o
                # formato esperado pelo provedor.
                break

        result.insert(0, {"role": "system", "content": addendum})
        return result

    def _build_client(self) -> OpenRouterClient:
        return OpenRouterClient(
            api_key=self.config.api_key,
            site_url=self.config.site_url,
            site_name=self.config.site_name,
        )

    @staticmethod
    def _is_image_model(model: str) -> bool:
        """Detecta modelos de geração/edição de imagem pelo padrão de ID
        usado no catálogo do OpenRouter (ex.: ``google/gemini-3.1-flash-
        image``, ``google/gemini-2.5-flash-image``).

        Desde o lançamento da Unified Image API do OpenRouter (final de
        junho de 2026), a chamada de chat completions só recebe a imagem de
        volta se o pedido incluir explicitamente ``modalities: ["image",
        "text"]``. Sem isso, o modelo tenta representar a imagem como uma
        sequência gigantesca de tokens de texto e a requisição estoura o
        limite de contexto do provedor (erro HTTP 400 "maximum context
        length"). É o mesmo passo que o site do OpenRouter faz quando você
        "ativa a ferramenta de geração de imagem" no chat deles.
        """
        return "-image" in model.lower()

    # ------------------------------------------------------------------ #
    # Cálculo automático de max_tokens                                     #
    # ------------------------------------------------------------------ #
    #
    # `max_tokens` limita a resposta (saída), mas conta dentro da mesma
    # janela de contexto do prompt (entrada). Um valor fixo alto (ex.:
    # herdado de um provedor com contexto enorme) estoura em modelos com
    # janela menor — foi exatamente o bug relatado com modelos ":free".
    #
    # A solução: consultar no catálogo do OpenRouter (`list_models`) o
    # `context_length` e o `top_provider.max_completion_tokens` do modelo
    # escolhido, estimar quantos tokens o prompt atual já ocupa, e usar o
    # que sobrar (com uma margem de segurança) como teto — sem nunca
    # ultrapassar o que o modelo de fato aceita.

    def _get_model_limits(
        self, client: OpenRouterClient, model: str
    ) -> Tuple[Optional[int], Optional[int]]:
        """Retorna (context_length, max_completion_tokens) do modelo, usando
        um cache de até 1h do catálogo pra não bater na API a cada envio.
        Se a consulta falhar (sem rede, chave inválida etc.), devolve
        (None, None) — o chamador cai de volta no comportamento antigo."""
        with self._model_cache_lock:
            expired = (time.monotonic() - self._model_cache_time) > self._MODEL_CACHE_TTL
            if expired or not self._model_cache:
                try:
                    raw = client.list_models()
                except OpenRouterError:
                    raw = None
                if raw is not None:
                    cache: Dict[str, Dict[str, Any]] = {}
                    for entry in raw:
                        if isinstance(entry, dict) and entry.get("id"):
                            cache[str(entry["id"])] = entry
                    self._model_cache = cache
                    self._model_cache_time = time.monotonic()
            entry = self._model_cache.get(model)
        return self._extract_limits(entry)

    @staticmethod
    def _extract_limits(
        entry: Optional[Dict[str, Any]],
    ) -> Tuple[Optional[int], Optional[int]]:
        if not isinstance(entry, dict):
            return None, None
        top_provider = entry.get("top_provider")
        top_provider = top_provider if isinstance(top_provider, dict) else {}

        context_length = entry.get("context_length")
        if context_length is None:
            context_length = top_provider.get("context_length")
        max_completion = top_provider.get("max_completion_tokens")

        def _to_int(value: Any) -> Optional[int]:
            try:
                parsed = int(value)
            except (TypeError, ValueError):
                return None
            return parsed if parsed > 0 else None

        return _to_int(context_length), _to_int(max_completion)

    @staticmethod
    def _estimate_prompt_tokens(messages: List[Any]) -> int:
        """Estimativa grosseira (mas conservadora) do tamanho do prompt em
        tokens, sem depender de um tokenizador específico de cada modelo."""
        total_chars = 0
        image_count = 0
        for message in messages:
            content = message.get("content") if isinstance(message, dict) else None
            if isinstance(content, str):
                total_chars += len(content)
            elif isinstance(content, list):
                for block in content:
                    if not isinstance(block, dict):
                        continue
                    if block.get("type") == "text":
                        total_chars += len(str(block.get("text", "")))
                    elif block.get("type") == "image_url":
                        image_count += 1
        # ~4 caracteres por token é uma aproximação razoável pra português/
        # inglês. Imagens custam uma quantidade de tokens que varia bastante
        # por modelo; usamos uma estimativa conservadora por imagem pra não
        # subestimar o gasto.
        return (total_chars // 4) + (image_count * 1500)

    def _resolve_max_tokens(
        self,
        client: OpenRouterClient,
        model: str,
        messages: List[Any],
    ) -> Optional[int]:
        configured = self.config.max_tokens  # None/0 em Preferências = automático
        context_length, max_completion = self._get_model_limits(client, model)

        limit_candidates: List[int] = []
        if max_completion:
            limit_candidates.append(max_completion)
        if context_length:
            prompt_estimate = self._estimate_prompt_tokens(messages)
            safety_margin = 256  # overhead de formatação, roles, etc.
            limit_candidates.append(max(context_length - prompt_estimate - safety_margin, 0))

        model_limit = min(limit_candidates) if limit_candidates else None

        if configured is None:
            # Modo automático: usa o maior valor seguro que o modelo aguenta.
            if model_limit is None:
                # Não conseguimos metadados do modelo (ex.: catálogo
                # indisponível): deixa o provedor aplicar seu próprio
                # padrão, em vez de arriscar um número inventado.
                return None
            return model_limit if model_limit >= 16 else None

        # Valor manual do usuário: nunca deixa passar do que o modelo aceita.
        if model_limit is not None:
            if model_limit < 16:
                return None
            return min(configured, model_limit)
        return configured

    # ------------------------------------------------------------------ #
    # Envio                                                                #
    # ------------------------------------------------------------------ #

    def send(
        self,
        model: str,
        messages: List[Any],
        on_delta: Callable[[str], None],
        # on_done(display, raw): `display` é o texto com os blocos de planilha
        # já trocados por links; `raw` é o texto original do modelo, com os
        # blocos intactos. São iguais quando não houve planilha na resposta.
        on_done: Callable[[str, str], None],
        on_error: Callable[[str], None],
        on_usage: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> bool:
        with self._lock:
            if self._inflight:
                return False
            self._inflight = True
            self._cancel.clear()

        threading.Thread(
            target=self._worker,
            args=(model, messages, on_delta, on_done, on_error, on_usage),
            daemon=True,
        ).start()
        return True

    def _worker(
        self,
        model: str,
        messages: List[Any],
        on_delta: Callable[[str], None],
        # on_done(display, raw): `display` é o texto com os blocos de planilha
        # já trocados por links; `raw` é o texto original do modelo, com os
        # blocos intactos. São iguais quando não houve planilha na resposta.
        on_done: Callable[[str, str], None],
        on_error: Callable[[str], None],
        on_usage: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> None:
        try:
            client = self._build_client()
            messages = self._inject_capabilities(messages)
            extra: Dict[str, Any] = {}
            if self._is_image_model(model):
                extra["modalities"] = ["image", "text"]
                # Não força max_tokens pra modelos de imagem: o valor
                # configurado em Preferências é pensado pra respostas de
                # texto e pode ser bem maior que a janela de contexto do
                # modelo de imagem, causando erro de limite de contexto
                # mesmo com a modalidade correta. Deixa o provedor decidir.
                max_tokens = None
            else:
                max_tokens = self._resolve_max_tokens(client, model, messages)

            if self.config.stream:
                full, images, usage = client.stream_chat(
                    model=model,
                    messages=messages,
                    on_delta=lambda text: GLib.idle_add(on_delta, text),
                    should_cancel=self._cancel.is_set,
                    temperature=self.config.temperature,
                    max_tokens=max_tokens,
                    **extra,
                )
            else:
                full, images, usage = client.chat(
                    model=model,
                    messages=messages,
                    temperature=self.config.temperature,
                    max_tokens=max_tokens,
                    **extra,
                )
                GLib.idle_add(on_delta, full)

            if images:
                markdown_links = self._save_generated_images(images)
                if markdown_links:
                    # Nome próprio: `extra` acima é o dict de kwargs da API.
                    suffix = ("\n\n" if full.strip() else "") + markdown_links
                    full += suffix
                    GLib.idle_add(on_delta, suffix)

            # Blocos ```ods / ```xlsx / ```csv viram arquivos locais + link
            # clicável. Roda depois do streaming, então a bolha ainda mostra
            # o bloco cru — o `on_done` substitui o texto pela versão
            # processada.
            #
            # O texto ANTES da conversão é devolvido junto: é o que permite a
            # janela reenviar a spec ao modelo num pedido de ajuste ("troca a
            # cor", "adiciona uma coluna"). Sem ele o histórico só teria o
            # link file://, e o modelo reinventaria a planilha do zero.
            display = full
            if full:
                display = ChatAssistant._save_generated_spreadsheets(full)

            if usage and on_usage is not None:
                GLib.idle_add(on_usage, usage)

            GLib.idle_add(on_done, display, full)
        except OpenRouterError as exc:
            GLib.idle_add(on_error, str(exc))
        except Exception as exc:  # pylint: disable=broad-except
            logger.exception("Erro inesperado no chat")
            GLib.idle_add(on_error, str(exc))
        finally:
            with self._lock:
                self._inflight = False

    # ------------------------------------------------------------------ #
    # Modelos                                                              #
    # ------------------------------------------------------------------ #

    def fetch_models(
        self,
        on_done: Callable[[List[str]], None],
        on_error: Callable[[str], None],
    ) -> None:
        def worker() -> None:
            try:
                client = self._build_client()
                raw = client.list_models()
                ids = sorted(
                    {
                        str(item.get("id")).strip()
                        for item in raw
                        if isinstance(item, dict) and item.get("id")
                    }
                )
                GLib.idle_add(on_done, ids)
            except OpenRouterError as exc:
                GLib.idle_add(on_error, str(exc))
            except Exception as exc:  # pylint: disable=broad-except
                logger.exception("Erro ao buscar modelos")
                GLib.idle_add(on_error, str(exc))

        threading.Thread(target=worker, daemon=True).start()

    # ------------------------------------------------------------------ #
    # Suporte a anexos                                                     #
    # ------------------------------------------------------------------ #

    @staticmethod
    def is_image(path: str) -> bool:
        return os.path.splitext(path)[1].lower() in _IMAGE_EXTENSIONS

    @staticmethod
    def supported_attachment(path: str) -> bool:
        ext = os.path.splitext(path)[1].lower()
        if ext in _TEXT_EXTENSIONS:
            return True
        if ext in _IMAGE_EXTENSIONS:
            return True
        if ext == ".pdf":
            return PDF_AVAILABLE
        if ext == ".odt":
            return ODT_AVAILABLE
        if ext == ".ods":
            return ODT_AVAILABLE
        if ext == ".xlsx":
            return XLSX_AVAILABLE
        return False

    @staticmethod
    def unsupported_reason(path: str) -> str:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".pdf" and not PDF_AVAILABLE:
            return "Instale python3-pypdf para anexar PDFs."
        if ext == ".odt" and not ODT_AVAILABLE:
            return "Instale python3-odfpy para anexar arquivos ODT."
        if ext == ".ods" and not ODT_AVAILABLE:
            return "Instale python3-odfpy para anexar arquivos ODS."
        if ext == ".xlsx" and not XLSX_AVAILABLE:
            return "Instale python3-openpyxl para anexar arquivos XLSX."
        if ext == ".docx":
            return "Arquivos .docx não são suportados. Converta para .odt ou .txt."
        return "Tipo de arquivo não suportado."

    @staticmethod
    def read_text_attachment(path: str) -> str:
        """Extrai texto de documentos (PDF, ODT, ODS, XLSX, texto puro)."""
        ext = os.path.splitext(path)[1].lower()
        if ext == ".pdf":
            return ChatAssistant._read_pdf(path)
        if ext == ".odt":
            return ChatAssistant._read_odt(path)
        if ext == ".ods":
            return ChatAssistant._read_ods(path)
        if ext == ".xlsx":
            return ChatAssistant._read_xlsx(path)
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
        except OSError as exc:
            raise RuntimeError(f"Erro ao ler arquivo: {exc}") from exc

    # Alias para compatibilidade com código existente.
    read_attachment = read_text_attachment

    # Assinaturas (magic numbers) dos formatos de imagem suportados. A
    # extensão do arquivo (.jpg, .png etc.) é só um nome escolhido pelo
    # usuário — nada garante que os bytes reais batem com ela. É comum uma
    # captura de tela sair como PNG e depois ser renomeada para .jpg (ou
    # vice-versa). Se a data URI declarar um MIME que não bate com os bytes
    # de fato, o provedor tenta decodificar no formato errado, falha e em
    # vez de dar erro claro alguns pipelines simplesmente repassam o blob
    # bruto ao modelo como texto — que aí "vê" só ruído/caracteres
    # repetidos. Por isso sempre inspecionamos os bytes primeiro.
    _MAGIC_SIGNATURES = (
        (b"\x89PNG\r\n\x1a\n", "image/png"),
        (b"\xff\xd8\xff", "image/jpeg"),
        (b"GIF87a", "image/gif"),
        (b"GIF89a", "image/gif"),
    )

    @staticmethod
    def _sniff_mime(data: bytes, fallback: str) -> str:
        for signature, mime in ChatAssistant._MAGIC_SIGNATURES:
            if data.startswith(signature):
                return mime
        # WEBP: RIFF....WEBP — os 4 primeiros bytes são "RIFF", só nos
        # bytes 8-12 é que aparece o marcador "WEBP" de fato.
        if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
            return "image/webp"
        # Formato não reconhecido pela assinatura: cai para o palpite feito
        # pela extensão do arquivo, em vez de travar o envio.
        return fallback

    @staticmethod
    def read_image_attachment(path: str) -> Dict[str, Any]:
        """
        Lê uma imagem e devolve o bloco multimodal para a API:
        {"type": "image_url", "image_url": {"url": "data:<mime>;base64,..."}}
        """
        guessed, _ = mimetypes.guess_type(path)
        if not guessed:
            ext = os.path.splitext(path)[1].lower().lstrip(".")
            # jpg → jpeg para compatibilidade com o padrão MIME
            ext = "jpeg" if ext == "jpg" else ext
            guessed = f"image/{ext}"
        try:
            with open(path, "rb") as f:
                raw = f.read()
        except OSError as exc:
            raise RuntimeError(f"Erro ao ler imagem: {exc}") from exc
        # A extensão é só um palpite; os bytes reais mandam.
        mime = ChatAssistant._sniff_mime(raw, guessed)
        data = base64.b64encode(raw).decode("ascii")
        return {
            "type": "image_url",
            "image_url": {"url": f"data:{mime};base64,{data}"},
        }

    _RE_DATA_URI = re.compile(r"^data:([^;]+);base64,(.+)$", re.DOTALL)

    # Blocos de planilha gerados pelo modelo.
    #
    # O group(1) é o formato (ods/xlsx) e o group(2) é o JSON — o formato
    # vem da cerca porque é assim que o modelo sinaliza a escolha pedida no
    # system prompt, sem precisar de mais uma chave dentro da spec.
    #
    # O `[ \t]*` tolera espaço em volta do nome da linguagem e o `\r?` cobre
    # quebra de linha estilo Windows. O `(?:```|\Z)` no fim aceita bloco sem
    # cerca de fechamento: modelos esquecem de fechar com frequência quando
    # o bloco é a última coisa da resposta, e sem essa tolerância o JSON
    # inteiro vazaria cru para a bolha.
    _RE_SHEET_BLOCK = re.compile(
        r"```[ \t]*(ods|xlsx)[ \t]*\r?\n(.*?)(?:```|\Z)", re.DOTALL | re.IGNORECASE
    )
    _RE_CSV_BLOCK = re.compile(
        r"```[ \t]*csv[ \t]*\r?\n(.*?)(?:```|\Z)", re.DOTALL | re.IGNORECASE
    )

    @staticmethod
    def _save_generated_images(image_urls: List[str]) -> str:
        """Salva imagens geradas pelo modelo (data URIs base64) em disco e
        devolve um bloco Markdown com links clicáveis para cada uma."""
        if not image_urls:
            return ""
        images_dir = os.path.join(GLib.get_user_data_dir(), "por-ai", "images")
        try:
            os.makedirs(images_dir, exist_ok=True)
        except OSError:
            return ""

        links: List[str] = []
        for url in image_urls:
            if not isinstance(url, str):
                continue
            match = ChatAssistant._RE_DATA_URI.match(url)
            if not match:
                if url.startswith(("http://", "https://")):
                    links.append(f"[Imagem gerada]({url})")
                continue
            mime, b64data = match.group(1), match.group(2)
            ext = mimetypes.guess_extension(mime) or ".png"
            if ext == ".jpe":
                ext = ".jpg"
            try:
                raw = base64.b64decode(b64data)
            except Exception:
                continue
            filename = f"{uuid.uuid4().hex}{ext}"
            path = os.path.join(images_dir, filename)
            try:
                with open(path, "wb") as f:
                    f.write(raw)
            except OSError:
                continue
            links.append(f"[Imagem gerada]({GLib.filename_to_uri(path, None)})")
        return "\n".join(links)

    @staticmethod
    def _spreadsheets_dir() -> Optional[str]:
        """Cria (se preciso) e devolve o diretório das planilhas geradas.
        Devolve None se não for possível criar — o chamador então deixa a
        resposta intacta."""
        path = os.path.join(GLib.get_user_data_dir(), "por-ai", "spreadsheets")
        try:
            os.makedirs(path, exist_ok=True)
        except OSError as exc:
            logger.warning("Não foi possível criar %s: %s", path, exc)
            return None
        return path

    @staticmethod
    def _save_generated_spreadsheets(text: str) -> str:
        """Converte blocos de planilha da resposta em arquivos locais.

        Três formatos são aceitos:

        * ```ods — JSON com dados e estilo, gera .ods formatado (cores,
          negrito, larguras). É o formato pedido no system prompt.
        * ```xlsx — mesma spec, saída em .xlsx. O system prompt só pede
          este quando o usuário menciona Excel explicitamente.
        * ```csv — texto puro, gera .csv sem formatação. Mantido porque
          nem todo modelo segue a instrução, e é melhor entregar uma
          planilha sem cor do que não entregar nada.

        Quando um bloco não pode ser convertido — JSON malformado, spec sem
        colunas, nenhum backend instalado, erro de escrita — o bloco
        original é preservado na resposta. Sumir com os dados do usuário
        por causa de uma falha de conversão seria pior do que mostrá-los
        crus.
        """
        directory = ChatAssistant._spreadsheets_dir()
        if directory is None:
            return text

        def _save_sheet(match: re.Match) -> str:
            formats = available_formats()
            if not formats:
                return match.group(0)

            fmt = match.group(1).lower()
            raw = match.group(2).strip()
            if not raw:
                return match.group(0)
            try:
                spec = json.loads(raw)
            except json.JSONDecodeError as exc:
                logger.info("Bloco de planilha com JSON inválido: %s", exc)
                return match.group(0)
            if not isinstance(spec, dict):
                return match.group(0)

            # Se o backend do formato pedido não existe, entrega no outro em
            # vez de devolver JSON cru: planilha no formato "errado" ainda é
            # uma planilha, e o usuário converte em dois cliques.
            if fmt not in formats:
                logger.info("Backend %s ausente, gerando em %s", fmt, formats[0])
                fmt = formats[0]

            path = os.path.join(directory, f"{uuid.uuid4().hex}.{fmt}")
            try:
                build_sheet(spec, path, fmt)
            except (SheetSpecError, OSError, ValueError) as exc:
                logger.info("Falha ao gerar planilha: %s", exc)
                return match.group(0)
            uri = GLib.filename_to_uri(path, None)
            return f"[Baixar planilha (.{fmt})]({uri})"

        def _save_csv(match: re.Match) -> str:
            content = match.group(1).strip()
            if not content:
                return match.group(0)
            path = os.path.join(directory, f"{uuid.uuid4().hex}.csv")
            try:
                with open(path, "w", encoding="utf-8", newline="") as handle:
                    handle.write(content)
            except OSError as exc:
                logger.info("Falha ao gravar planilha csv: %s", exc)
                return match.group(0)
            uri = GLib.filename_to_uri(path, None)
            return f"[Baixar planilha (.csv)]({uri})"

        text = ChatAssistant._RE_SHEET_BLOCK.sub(_save_sheet, text)
        return ChatAssistant._RE_CSV_BLOCK.sub(_save_csv, text)

    # ------------------------------------------------------------------ #
    # Leitores específicos                                                 #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _read_pdf(path: str) -> str:
        if not PDF_AVAILABLE:
            raise RuntimeError("Instale python3-pypdf para ler PDFs.")
        try:
            reader = PdfReader(path)
            parts = [p.extract_text() for p in reader.pages if p.extract_text()]
            text = "\n".join(parts)
        except Exception as exc:
            raise RuntimeError(f"Erro ao ler PDF: {exc}") from exc
        if not text.strip():
            raise RuntimeError(
                "Não foi possível extrair texto do PDF "
                "(pode ser imagem ou estar vazio)."
            )
        return text

    @staticmethod
    def _read_odt(path: str) -> str:
        if not ODT_AVAILABLE:
            raise RuntimeError("Instale python3-odfpy para ler ODT.")
        try:
            doc = odf_load(path)
            paragraphs = doc.getElementsByType(P)
            lines = []
            for para in paragraphs:
                text = "".join(
                    node.data
                    for node in para.childNodes
                    if node.nodeType == node.TEXT_NODE
                )
                lines.append(text)
            return "\n".join(lines)
        except Exception as exc:
            raise RuntimeError(f"Erro ao ler ODT: {exc}") from exc

    @staticmethod
    def _read_ods(path: str) -> str:
        if not ODT_AVAILABLE:
            raise RuntimeError("Instale python3-odfpy para ler arquivos ODS.")
        try:
            from odf.table import Table, TableRow, TableCell
            doc = odf_load(path)
            lines = []
            for table in doc.getElementsByType(Table):
                for row in table.getElementsByType(TableRow):
                    cells = []
                    for cell in row.getElementsByType(TableCell):
                        text = "".join(
                            node.data
                            for node in cell.childNodes
                            if node.nodeType == node.TEXT_NODE
                        )
                        cells.append(text)
                    lines.append("\t".join(cells))
            return "\n".join(lines)
        except Exception as exc:
            raise RuntimeError(f"Erro ao ler ODS: {exc}") from exc

    @staticmethod
    def _read_xlsx(path: str) -> str:
        """Lê uma planilha .xlsx como texto tabulado, uma linha por linha da
        planilha. ``data_only=True`` traz o resultado das fórmulas em vez da
        fórmula em si — é o que interessa ao modelo."""
        if not XLSX_AVAILABLE:
            raise RuntimeError("Instale python3-openpyxl para ler arquivos XLSX.")
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
            blocks = []
            for sheet in workbook.worksheets:
                lines = []
                for row in sheet.iter_rows(values_only=True):
                    if row is None:
                        continue
                    cells = ["" if value is None else str(value) for value in row]
                    if any(cell.strip() for cell in cells):
                        lines.append("\t".join(cells))
                if lines:
                    blocks.append(f"# {sheet.title}\n" + "\n".join(lines))
            workbook.close()
        except Exception as exc:
            raise RuntimeError(f"Erro ao ler XLSX: {exc}") from exc
        if not blocks:
            raise RuntimeError("A planilha está vazia.")
        return "\n\n".join(blocks)
